import pandas as pd
import numpy as np
import io
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class TournamentManager:
    def __init__(self):
        """Initialize the tournament manager with empty dataframes for teams, matches, and standings."""
        self.teams = pd.DataFrame(columns=["Nome squadra", "Referente", "Contatto", "Quota pagata", "Note"])
        self.groups = {}  # Dictionary to store team assignments to groups
        self.matches = {}  # Dictionary to store match schedules by group
        self.playoffs = pd.DataFrame(columns=["Fase", "Squadra 1", "Squadra 2", "Set Squadra 1", "Set Squadra 2", "Vincitore"])
        self.final_standings = pd.DataFrame(columns=["Posizione", "Nome squadra"])
        
        # Tournament settings
        self.points_win = 3
        self.points_loss = 0
        self.points_tiebreak = 1  # Points for losing in a tiebreak (e.g., 2-3)
    
    def add_team(self, name, contact_person, contact_info, payment_status, notes):
        """
        Add a new team to the tournament.
        
        Args:
            name: Team name
            contact_person: Team representative's name
            contact_info: Contact information for the team
            payment_status: Whether the team has paid the registration fee
            notes: Additional notes
            
        Returns:
            bool: True if team was added successfully, False if team already exists
        """
        # Check if team already exists
        if name in self.teams["Nome squadra"].values:
            return False
        
        # Add the team
        new_team = pd.DataFrame({
            "Nome squadra": [name],
            "Referente": [contact_person],
            "Contatto": [contact_info],
            "Quota pagata": [payment_status],
            "Note": [notes]
        })
        
        self.teams = pd.concat([self.teams, new_team], ignore_index=True)
        return True
    
    def create_groups(self, num_groups, teams_per_group):
        """
        Create tournament groups by distributing teams.
        
        Args:
            num_groups: Number of groups to create
            teams_per_group: Number of teams per group
        """
        # Reset existing groups
        self.groups = {}
        self.matches = {}
        
        # Get team names
        team_names = self.teams["Nome squadra"].tolist()
        
        # Shuffle teams to randomize group assignments
        np.random.shuffle(team_names)
        
        # Create groups (A, B, C, etc.)
        group_names = [chr(65 + i) for i in range(num_groups)]  # A, B, C, ...
        
        # Assign teams to groups
        for i, group in enumerate(group_names):
            start_idx = i * teams_per_group
            end_idx = start_idx + teams_per_group
            
            if start_idx < len(team_names):
                self.groups[group] = team_names[start_idx:min(end_idx, len(team_names))]
    
    def generate_matches(self):
        """Generate round-robin match schedules for each group."""
        self.matches = {}
        
        for group, teams in self.groups.items():
            matches = []
            
            # Generate round-robin schedule
            for i in range(len(teams)):
                for j in range(i + 1, len(teams)):
                    matches.append({
                        "Squadra 1": teams[i],
                        "Squadra 2": teams[j],
                        "Set Squadra 1": None,
                        "Set Squadra 2": None,
                        "Vincitore": ""
                    })
            
            self.matches[group] = pd.DataFrame(matches)
    
    def update_match_results(self, group, updated_df):
        """
        Update match results and recalculate winners.
        
        Args:
            group: Group identifier (A, B, C, etc.)
            updated_df: DataFrame with updated match results
        """
        # Update match data
        self.matches[group] = updated_df.copy()
        
        # Calculate winners
        for idx, row in self.matches[group].iterrows():
            team1_sets = row["Set Squadra 1"]
            team2_sets = row["Set Squadra 2"]
            
            # Skip if results not entered yet
            if pd.isna(team1_sets) or pd.isna(team2_sets):
                self.matches[group].at[idx, "Vincitore"] = ""
                continue
            
            # Determine winner
            if team1_sets > team2_sets:
                self.matches[group].at[idx, "Vincitore"] = row["Squadra 1"]
            elif team2_sets > team1_sets:
                self.matches[group].at[idx, "Vincitore"] = row["Squadra 2"]
            else:
                self.matches[group].at[idx, "Vincitore"] = "Draw"
    
    def calculate_group_standings(self, group):
        """
        Calculate standings for a specific group.
        
        Args:
            group: Group identifier (A, B, C, etc.)
            
        Returns:
            DataFrame: Group standings with teams, matches, sets, and points
        """
        if group not in self.matches:
            return pd.DataFrame()
        
        matches_df = self.matches[group]
        teams = self.groups[group]
        
        # Initialize standings dataframe
        standings = pd.DataFrame({
            "Team": teams,
            "Matches Played": 0,
            "Wins": 0,
            "Losses": 0,
            "Sets Won": 0,
            "Sets Lost": 0,
            "Points": 0
        })
        
        # Process each match
        for _, match in matches_df.iterrows():
            team1 = match["Squadra 1"]
            team2 = match["Squadra 2"]
            sets1 = match["Set Squadra 1"]
            sets2 = match["Set Squadra 2"]
            
            # Skip if match results not entered
            if pd.isna(sets1) or pd.isna(sets2):
                continue
            
            # Update team 1 statistics
            team1_idx = standings[standings["Team"] == team1].index[0]
            standings.at[team1_idx, "Matches Played"] += 1
            standings.at[team1_idx, "Sets Won"] += sets1
            standings.at[team1_idx, "Sets Lost"] += sets2
            
            # Update team 2 statistics
            team2_idx = standings[standings["Team"] == team2].index[0]
            standings.at[team2_idx, "Matches Played"] += 1
            standings.at[team2_idx, "Sets Won"] += sets2
            standings.at[team2_idx, "Sets Lost"] += sets1
            
            # Update wins, losses, and points
            if sets1 > sets2:
                standings.at[team1_idx, "Wins"] += 1
                standings.at[team2_idx, "Losses"] += 1
                
                # Award points (3 for win, 0 for loss, 1 for tiebreak loss)
                standings.at[team1_idx, "Points"] += self.points_win
                
                # Check for tiebreak loss (e.g., 3-2)
                if sets2 == sets1 - 1 and sets1 == 3:
                    standings.at[team2_idx, "Points"] += self.points_tiebreak
                else:
                    standings.at[team2_idx, "Points"] += self.points_loss
                    
            elif sets2 > sets1:
                standings.at[team2_idx, "Wins"] += 1
                standings.at[team1_idx, "Losses"] += 1
                
                # Award points
                standings.at[team2_idx, "Points"] += self.points_win
                
                # Check for tiebreak loss
                if sets1 == sets2 - 1 and sets2 == 3:
                    standings.at[team1_idx, "Points"] += self.points_tiebreak
                else:
                    standings.at[team1_idx, "Points"] += self.points_loss
        
        # Sort standings by points (descending), then by set difference
        standings["Set Difference"] = standings["Sets Won"] - standings["Sets Lost"]
        standings = standings.sort_values(by=["Points", "Set Difference", "Sets Won"], ascending=False)
        
        # Drop the set difference column (used only for sorting)
        standings = standings.drop(columns=["Set Difference"])
        
        return standings
    
    def check_groups_complete(self):
        """
        Check if all group matches have been completed.
        
        Returns:
            bool: True if all matches have results, False otherwise
        """
        if not self.matches:
            return False
        
        for group, matches_df in self.matches.items():
            if matches_df["Set Squadra 1"].isna().any() or matches_df["Set Squadra 2"].isna().any():
                return False
        
        return True
    
    def generate_playoffs(self, teams_advancing):
        """
        Generate playoff brackets based on group standings.
        
        Args:
            teams_advancing: Number of teams advancing from each group
        """
        phases = []
        qualified_teams = []
        
        # Get top teams from each group
        for group in self.groups.keys():
            standings = self.calculate_group_standings(group)
            qualified_teams.extend([(group, team, pos) for pos, team in enumerate(standings["Team"].head(teams_advancing))])
        
        total_teams = len(qualified_teams)
        
        # Calculate playoff structure
        if total_teams >= 8:
            # Quarterfinals
            num_quarters = 4
            for i in range(num_quarters):
                # Determine teams for quarterfinal i
                # This is a simplistic approach, ideally you'd want to distribute by group and position
                team1_idx = i
                team2_idx = total_teams - 1 - i
                
                if team1_idx < len(qualified_teams) and team2_idx < len(qualified_teams):
                    team1 = qualified_teams[team1_idx][1]
                    team2 = qualified_teams[team2_idx][1]
                    
                    phases.append({
                        "Fase": "Quarterfinals",
                        "Squadra 1": team1,
                        "Squadra 2": team2,
                        "Set Squadra 1": None,
                        "Set Squadra 2": None,
                        "Vincitore": ""
                    })
            
            # Semifinals
            phases.append({
                "Fase": "Semifinals",
                "Squadra 1": "QF1 Winner",
                "Squadra 2": "QF2 Winner",
                "Set Squadra 1": None,
                "Set Squadra 2": None,
                "Vincitore": ""
            })
            
            phases.append({
                "Fase": "Semifinals",
                "Squadra 1": "QF3 Winner",
                "Squadra 2": "QF4 Winner",
                "Set Squadra 1": None,
                "Set Squadra 2": None,
                "Vincitore": ""
            })
            
            # Finals
            phases.append({
                "Fase": "Finals",
                "Squadra 1": "SF1 Winner",
                "Squadra 2": "SF2 Winner",
                "Set Squadra 1": None,
                "Set Squadra 2": None,
                "Vincitore": ""
            })
            
            # Third place match
            phases.append({
                "Fase": "Third Place",
                "Squadra 1": "SF1 Loser",
                "Squadra 2": "SF2 Loser",
                "Set Squadra 1": None,
                "Set Squadra 2": None,
                "Vincitore": ""
            })
            
        elif total_teams >= 4:
            # Semifinals only
            for i in range(2):
                team1_idx = i
                team2_idx = total_teams - 1 - i
                
                if team1_idx < len(qualified_teams) and team2_idx < len(qualified_teams):
                    team1 = qualified_teams[team1_idx][1]
                    team2 = qualified_teams[team2_idx][1]
                    
                    phases.append({
                        "Fase": "Semifinals",
                        "Squadra 1": team1,
                        "Squadra 2": team2,
                        "Set Squadra 1": None,
                        "Set Squadra 2": None,
                        "Vincitore": ""
                    })
            
            # Finals
            phases.append({
                "Fase": "Finals",
                "Squadra 1": "SF1 Winner",
                "Squadra 2": "SF2 Winner",
                "Set Squadra 1": None,
                "Set Squadra 2": None,
                "Vincitore": ""
            })
            
            # Third place match
            phases.append({
                "Fase": "Third Place",
                "Squadra 1": "SF1 Loser",
                "Squadra 2": "SF2 Loser",
                "Set Squadra 1": None,
                "Set Squadra 2": None,
                "Vincitore": ""
            })
            
        else:
            # Just finals
            if len(qualified_teams) >= 2:
                phases.append({
                    "Fase": "Finals",
                    "Squadra 1": qualified_teams[0][1],
                    "Squadra 2": qualified_teams[1][1],
                    "Set Squadra 1": None,
                    "Set Squadra 2": None,
                    "Vincitore": ""
                })
        
        self.playoffs = pd.DataFrame(phases)
    
    def update_playoff_results(self, updated_df):
        """
        Update playoff match results and propagate winners to next rounds.
        
        Args:
            updated_df: DataFrame with updated playoff match results
        """
        self.playoffs = updated_df.copy()
        
        # Calculate winners and update next round matches
        for idx, row in self.playoffs.iterrows():
            team1_sets = row["Set Squadra 1"]
            team2_sets = row["Set Squadra 2"]
            phase = row["Fase"]
            
            # Skip if results not entered yet
            if pd.isna(team1_sets) or pd.isna(team2_sets):
                self.playoffs.at[idx, "Vincitore"] = ""
                continue
            
            # Determine winner
            winner = ""
            loser = ""
            if team1_sets > team2_sets:
                winner = row["Squadra 1"]
                loser = row["Squadra 2"]
            elif team2_sets > team1_sets:
                winner = row["Squadra 2"]
                loser = row["Squadra 1"]
            
            self.playoffs.at[idx, "Vincitore"] = winner
            
            # Propagate winners to next round
            if phase == "Quarterfinals":
                qf_number = idx + 1  # QF1, QF2, etc.
                
                # Find semifinals that reference this quarterfinal
                for sf_idx, sf_row in self.playoffs[self.playoffs["Fase"] == "Semifinals"].iterrows():
                    if f"QF{qf_number} Winner" in [sf_row["Squadra 1"], sf_row["Squadra 2"]]:
                        if sf_row["Squadra 1"] == f"QF{qf_number} Winner":
                            self.playoffs.at[sf_idx, "Squadra 1"] = winner
                        else:
                            self.playoffs.at[sf_idx, "Squadra 2"] = winner
            
            elif phase == "Semifinals":
                sf_number = 1 if idx == min(self.playoffs[self.playoffs["Fase"] == "Semifinals"].index) else 2
                
                # Update finals
                for final_idx, final_row in self.playoffs[self.playoffs["Fase"] == "Finals"].iterrows():
                    if f"SF{sf_number} Winner" in [final_row["Squadra 1"], final_row["Squadra 2"]]:
                        if final_row["Squadra 1"] == f"SF{sf_number} Winner":
                            self.playoffs.at[final_idx, "Squadra 1"] = winner
                        else:
                            self.playoffs.at[final_idx, "Squadra 2"] = winner
                
                # Update third place match
                for third_idx, third_row in self.playoffs[self.playoffs["Fase"] == "Third Place"].iterrows():
                    if f"SF{sf_number} Loser" in [third_row["Squadra 1"], third_row["Squadra 2"]]:
                        if third_row["Squadra 1"] == f"SF{sf_number} Loser":
                            self.playoffs.at[third_idx, "Squadra 1"] = loser
                        else:
                            self.playoffs.at[third_idx, "Squadra 2"] = loser
    
    def check_playoffs_complete(self):
        """
        Check if all playoff matches have been completed.
        
        Returns:
            bool: True if all matches have results, False otherwise
        """
        if self.playoffs.empty:
            return False
        
        return not (self.playoffs["Set Squadra 1"].isna().any() or self.playoffs["Set Squadra 2"].isna().any())
    
    def generate_final_standings(self):
        """Generate final tournament standings based on playoff results."""
        positions = []
        
        # Get the playoff results
        if self.playoffs.empty or not self.check_playoffs_complete():
            return
        
        # Find the winner and runner-up from Finals
        finals = self.playoffs[self.playoffs["Fase"] == "Finals"].iloc[0]
        winner = finals["Vincitore"]
        runner_up = finals["Squadra 1"] if finals["Vincitore"] == finals["Squadra 2"] else finals["Squadra 2"]
        
        positions.append({"Posizione": 1, "Nome squadra": winner})
        positions.append({"Posizione": 2, "Nome squadra": runner_up})
        
        # Find third place winner
        third_place_match = self.playoffs[self.playoffs["Fase"] == "Third Place"]
        if not third_place_match.empty:
            third_place = third_place_match.iloc[0]["Vincitore"]
            fourth_place = third_place_match.iloc[0]["Squadra 1"] if third_place_match.iloc[0]["Vincitore"] == third_place_match.iloc[0]["Squadra 2"] else third_place_match.iloc[0]["Squadra 2"]
            
            positions.append({"Posizione": 3, "Nome squadra": third_place})
            positions.append({"Posizione": 4, "Nome squadra": fourth_place})
        
        # Get the quarterfinalists (5th-8th place)
        quarterfinals = self.playoffs[self.playoffs["Fase"] == "Quarterfinals"]
        if not quarterfinals.empty:
            qf_losers = []
            for _, qf in quarterfinals.iterrows():
                loser = qf["Squadra 1"] if qf["Vincitore"] == qf["Squadra 2"] else qf["Squadra 2"]
                qf_losers.append(loser)
            
            # Assign positions 5-8 (or equivalent)
            for i, loser in enumerate(qf_losers):
                positions.append({"Posizione": 5 + i, "Nome squadra": loser})
        
        # Add the remaining teams from group stage
        pos = len(positions) + 1
        for group in self.groups.keys():
            standings = self.calculate_group_standings(group)
            
            # Skip teams that already appear in playoffs
            for _, team in enumerate(standings["Team"]):
                if not any(p["Nome squadra"] == team for p in positions):
                    positions.append({"Posizione": pos, "Nome squadra": team})
                    pos += 1
        
        self.final_standings = pd.DataFrame(positions)
    
    def export_to_excel(self):
        """
        Export all tournament data to an Excel file.
        
        Returns:
            BytesIO: Excel file as bytes
        """
        output = io.BytesIO()
        
        # Create workbook
        wb = Workbook()
        
        # Create registration sheet
        ws_reg = wb.active
        ws_reg.title = "Iscrizioni"
        
        # Helper function to write dataframe to worksheet
        def write_df_to_sheet(sheet, df):
            for r in dataframe_to_rows(df, index=False, header=True):
                sheet.append(r)
            
            # Format header row
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
        
        # Write registrations
        write_df_to_sheet(ws_reg, self.teams)
        
        # Write groups and matches
        for group, teams in self.groups.items():
            # Create group sheet
            ws_group = wb.create_sheet(f"Girone {group}")
            
            # Write team list
            teams_df = pd.DataFrame({"Squadre": teams})
            write_df_to_sheet(ws_group, teams_df)
            
            # Add separator row
            ws_group.append([])
            
            # Write matches
            if group in self.matches:
                start_row = ws_group.max_row + 1
                write_df_to_sheet(ws_group, self.matches[group])
            
            # Add separator row
            ws_group.append([])
            
            # Write standings
            standings = self.calculate_group_standings(group)
            if not standings.empty:
                start_row = ws_group.max_row + 1
                ws_group.append(["Classifica"])
                ws_group.cell(row=start_row, column=1).font = Font(bold=True)
                write_df_to_sheet(ws_group, standings)
        
        # Write playoffs
        if not self.playoffs.empty:
            ws_playoffs = wb.create_sheet("Fasi Finali")
            write_df_to_sheet(ws_playoffs, self.playoffs)
        
        # Write final standings
        if not self.final_standings.empty:
            ws_standings = wb.create_sheet("Classifica Finale")
            write_df_to_sheet(ws_standings, self.final_standings)
        
        # Save to BytesIO
        wb.save(output)
        output.seek(0)
        
        return output

# Streamlit UI
def main():
    st.set_page_config(page_title="Tournament Manager", layout="wide")
    st.title("🏐 Tournament Manager")
    
    # Initialize session state
    if 'tournament' not in st.session_state:
        st.session_state.tournament = TournamentManager()
    
    # Create tabs for different sections
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Team Registration", 
        "Group Stage", 
        "Playoffs", 
        "Final Standings", 
        "Export Data"
    ])
    
    with tab1:
        st.header("Team Registration")
        
        with st.expander("Add New Team"):
            with st.form("team_form"):
                col1, col2 = st.columns(2)
                name = col1.text_input("Team Name*")
                contact_person = col1.text_input("Contact Person*")
                contact_info = col2.text_input("Contact Info*")
                payment_status = col2.selectbox("Payment Status", ["Yes", "No"])
                notes = st.text_area("Notes")
                
                if st.form_submit_button("Add Team"):
                    if name and contact_person and contact_info:
                        added = st.session_state.tournament.add_team(
                            name, contact_person, contact_info, payment_status, notes
                        )
                        if added:
                            st.success("Team added successfully!")
                        else:
                            st.error("Team with this name already exists.")
                    else:
                        st.warning("Please fill all required fields (*)")
        
        st.subheader("Registered Teams")
        if not st.session_state.tournament.teams.empty:
            st.dataframe(st.session_state.tournament.teams)
        else:
            st.info("No teams registered yet.")
    
    with tab2:
        st.header("Group Stage Management")
        
        if not st.session_state.tournament.teams.empty:
            with st.expander("Create Groups"):
                col1, col2 = st.columns(2)
                num_groups = col1.number_input("Number of Groups", min_value=1, max_value=10, value=2)
                teams_per_group = col2.number_input("Teams per Group", min_value=2, max_value=10, value=4)
                
                if st.button("Generate Groups"):
                    st.session_state.tournament.create_groups(num_groups, teams_per_group)
                    st.session_state.tournament.generate_matches()
                    st.success("Groups created successfully!")
            
            if st.session_state.tournament.groups:
                st.subheader("Groups")
                for group, teams in st.session_state.tournament.groups.items():
                    with st.expander(f"Group {group}"):
                        st.write(f"Teams in Group {group}:")
                        st.dataframe(pd.DataFrame(teams, columns=["Team"]))
                        
                        if group in st.session_state.tournament.matches:
                            st.write("Matches:")
                            edited_df = st.data_editor(
                                st.session_state.tournament.matches[group],
                                key=f"matches_{group}"
                            )
                            
                            if st.button(f"Update Results for Group {group}"):
                                st.session_state.tournament.update_match_results(group, edited_df)
                                st.success("Results updated!")
                            
                            st.subheader(f"Group {group} Standings")
                            standings = st.session_state.tournament.calculate_group_standings(group)
                            st.dataframe(standings)
        else:
            st.info("Please register teams first and create groups.")
    
    with tab3:
        st.header("Playoff Stage")
        
        if st.session_state.tournament.groups:
            if st.session_state.tournament.check_groups_complete():
                with st.expander("Generate Playoff Bracket"):
                    teams_advancing = st.number_input(
                        "Teams Advancing from Each Group", 
                        min_value=1, 
                        max_value=4,
                        value=2
                    )
                    
                    if st.button("Generate Playoff Matches"):
                        st.session_state.tournament.generate_playoffs(teams_advancing)
                        st.success("Playoff bracket generated!")
                
                if not st.session_state.tournament.playoffs.empty:
                    st.subheader("Playoff Matches")
                    edited_playoffs = st.data_editor(
                        st.session_state.tournament.playoffs,
                        key="playoffs_editor"
                    )
                    
                    if st.button("Update Playoff Results"):
                        st.session_state.tournament.update_playoff_results(edited_playoffs)
                        st.success("Playoff results updated!")
            else:
                st.warning("Please complete all group stage matches first.")
        else:
            st.info("Please create groups and complete group stage first.")
    
    with tab4:
        st.header("Final Standings")
        
        if not st.session_state.tournament.playoffs.empty:
            if st.session_state.tournament.check_playoffs_complete():
                if st.button("Generate Final Standings"):
                    st.session_state.tournament.generate_final_standings()
                
                if not st.session_state.tournament.final_standings.empty:
                    st.subheader("Tournament Final Standings")
                    st.dataframe(st.session_state.tournament.final_standings)
            else:
                st.warning("Please complete all playoff matches first.")
        else:
            st.info("Final standings will be available after playoffs are completed.")
    
    with tab5:
        st.header("Export Tournament Data")
        
        if not st.session_state.tournament.teams.empty:
            excel_data = st.session_state.tournament.export_to_excel()
            st.download_button(
                label="Download Excel File",
                data=excel_data,
                file_name="tournament_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("No data to export yet. Please register teams first.")

if __name__ == "__main__":
    main()